import os
import json
import math
import shutil
import httpx
from dotenv import load_dotenv
load_dotenv()
from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import HTMLResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
import uvicorn
import ezdxf
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from takeoff_analysis import analyze_dxf_json, clean_text, analyze_sheet_bbox, extract_dxf_data_for_ai, extract_height_texts_for_ai

def sanitize_surrogates(obj: Any) -> Any:
    """
    재귀적으로 모든 문자열에서 surrogate 문자를 제거하여 JSON 직렬화 가능하게 만듭니다.
    """
    if isinstance(obj, str):
        # surrogate 문자 제거
        return obj.encode('utf-8', 'surrogatepass').decode('utf-8', 'ignore')
    elif isinstance(obj, dict):
        return {k: sanitize_surrogates(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [sanitize_surrogates(item) for item in obj]
    elif isinstance(obj, tuple):
        return tuple(sanitize_surrogates(item) for item in obj)
    else:
        return obj

app = FastAPI(title="H-Beam Takeoff Verification System API")

OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY")
OPENROUTER_MODEL = os.getenv("OPENROUTER_MODEL", "google/gemini-2.5-flash")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# API 1: DXF 파일 업로드 및 분석 데이터 반환
@app.post("/api/analyze")
async def analyze_file(file: UploadFile = File(...)):
    # 파일 확장자 검사
    if not file.filename.endswith(".dxf"):
        raise HTTPException(status_code=400, detail="Only DXF files are allowed.")
    
    file_path = os.path.join(DATA_DIR, file.filename)
    
    # 디스크에 저장
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
        
    # 분석 실행
    try:
        result = analyze_dxf_json(file_path)
        if "error" in result:
            raise HTTPException(status_code=500, detail=result["error"])
    except Exception as e:
        import traceback
        err_msg = traceback.format_exc()
        print(f"Error during analysis: {err_msg}")
        raise HTTPException(status_code=500, detail=f"Analysis crashed: {str(e)}\n{err_msg}")
    
    # surrogate 문자 제거 (JSON 직렬화 오류 방지)
    result = sanitize_surrogates(result)
        
    # 메타데이터 파일 확인 및 자동 로드
    meta_path = file_path + ".meta.json"
    meta_data = None
    if os.path.exists(meta_path):
        try:
            with open(meta_path, "r", encoding="utf-8") as f:
                meta_data = json.load(f)
                meta_data = sanitize_surrogates(meta_data)
        except Exception as e:
            print(f"Failed to read meta file: {e}")
            
    return {
        "filename": file.filename,
        "analysis": result,
        "metadata": meta_data
    }

# API 2: 메타데이터 저장
class SaveMetaRequest(BaseModel):
    filename: str
    metadata: dict

@app.post("/api/save_meta")
async def save_meta(request: SaveMetaRequest):
    file_path = os.path.join(DATA_DIR, request.filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Target DXF file not found.")
        
    meta_path = file_path + ".meta.json"
    try:
        with open(meta_path, "w", encoding="utf-8") as f:
            json.dump(request.metadata, f, ensure_ascii=False, indent=4)
        return {"status": "success", "message": "Metadata saved successfully."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to save metadata: {str(e)}")


# API: 수동 추가된 시트 영역의 도면명 자동 추출
class ExtractSheetNameRequest(BaseModel):
    filename: str
    bbox: List[float]

@app.post("/api/extract_sheet_name")
async def extract_sheet_name(request: ExtractSheetNameRequest):
    file_path = os.path.join(DATA_DIR, request.filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Target DXF file not found.")
    
    try:
        from takeoff_analysis import extract_sheet_name_from_bbox
        result = extract_sheet_name_from_bbox(file_path, request.bbox)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to extract sheet name: {str(e)}")


# API: 수동 추가된 시트 영역에 대한 부분 재분석 실행
class AnalyzeSheetZoneRequest(BaseModel):
    filename: str
    sheet_id: str
    bbox: List[float]
    scale: float

@app.post("/api/analyze_sheet_zone")
async def analyze_sheet_zone(request: AnalyzeSheetZoneRequest):
    file_path = os.path.join(DATA_DIR, request.filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Target DXF file not found.")
    
    try:
        result = analyze_sheet_bbox(file_path, request.bbox, request.scale, request.sheet_id)
        if "error" in result:
            raise HTTPException(status_code=500, detail=result["error"])
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to analyze sheet zone: {str(e)}")


# API 3: 최종 승인 및 엑셀/마킹 CAD 다운로드
class ExportRequest(BaseModel):
    filename: str
    sheets: List[dict] # 사용자 승인된 시트 목록
    columns: List[dict] # 최종 승인된 기둥 데이터
    beams: List[dict] # 최종 승인된 보 데이터

@app.post("/api/export")
async def export_takeoff(request: ExportRequest):
    dxf_path = os.path.join(DATA_DIR, request.filename)
    if not os.path.exists(dxf_path):
        raise HTTPException(status_code=404, detail="Original DXF file not found.")

    # 전체 메타데이터 파일 로드하여 참조 시트용 원본 데이터 확보 + source 정보 보강
    all_sheets_dict = {}
    meta_path = dxf_path + ".meta.json"
    meta_cols_by_id = {}
    meta_beams_by_id = {}
    if os.path.exists(meta_path):
        try:
            with open(meta_path, "r", encoding="utf-8") as f:
                meta_data = json.load(f)
                all_sheets_dict = {s["id"]: s for s in meta_data.get("sheets", [])}
                # id 기반 source 조회용 딕셔너리 생성
                meta_cols_by_id = {c["id"]: c for c in meta_data.get("columns", []) if "id" in c}
                meta_beams_by_id = {b["id"]: b for b in meta_data.get("beams", []) if "id" in b}
        except Exception as e:
            print(f"Failed to load meta file in export: {e}")

    # 디버그: 메타데이터 기둥 수 및 source 있는 기둥 수 확인
    print(f"[DEBUG] 메타데이터 기둥 수: {len(meta_cols_by_id)}")
    print(f"[DEBUG] 메타데이터에 source 있는 기둥: {sum(1 for c in meta_cols_by_id.values() if c.get('source'))}")

    # request.columns/beams에 source 필드가 없는 경우 메타데이터에서 보강
    for col in request.columns:
        if not col.get("source"):
            col_id = col.get("id", "")
            meta_col = meta_cols_by_id.get(col_id)
            if meta_col and meta_col.get("source"):
                col["source"] = meta_col["source"]
            else:
                # id 매칭 실패 시 좌표 근사 매칭으로 source 복원
                for mc in meta_cols_by_id.values():
                    if mc.get("source") and mc.get("sheet_id") == col.get("sheet_id"):
                        if abs(mc.get("cx", 0) - col.get("cx", 0)) < 1500 and abs(mc.get("cy", 0) - col.get("cy", 0)) < 1500:
                            col["source"] = mc["source"]
                            break
        # height 필드가 없거나 0인 경우 메타데이터에서 보강
        if not col.get("height"):
            col_id = col.get("id", "")
            meta_col = meta_cols_by_id.get(col_id)
            if meta_col and meta_col.get("height"):
                col["height"] = meta_col["height"]

    for beam in request.beams:
        if not beam.get("source"):
            beam_id = beam.get("id", "")
            meta_beam = meta_beams_by_id.get(beam_id)
            if meta_beam and meta_beam.get("source"):
                beam["source"] = meta_beam["source"]
            else:
                # id 매칭 실패 시 좌표 근사 매칭으로 source 복원
                for mb in meta_beams_by_id.values():
                    if mb.get("source") and mb.get("sheet_id") == beam.get("sheet_id"):
                        b_start = beam.get("start", [0, 0])
                        mb_start = mb.get("start", [0, 0])
                        if abs(mb_start[0] - b_start[0]) < 2000 and abs(mb_start[1] - b_start[1]) < 2000:
                            beam["source"] = mb["source"]
                            break

    # 디버그: AI 기둥/보 수 로그 출력
    ai_col_cnt = sum(1 for c in request.columns if c.get("source") == "ai")
    ai_beam_cnt = sum(1 for b in request.beams if b.get("source") == "ai")
    both_beam_cnt = sum(1 for b in request.beams if b.get("source") == "both")
    print(f"[EXPORT] 전체 기둥: {len(request.columns)}개, AI 기둥: {ai_col_cnt}개")
    print(f"[EXPORT] 전체 보: {len(request.beams)}개, AI 보: {ai_beam_cnt}개, Both 보: {both_beam_cnt}개")
    
    # 1. DXF 마킹 렌더링 파일 생성 (생성 중지)
    marked_dxf_filename = None

    # 2. 엑셀 적산 내역서 생성
    excel_filename = request.filename.replace(".dxf", "_takeoff.xlsx")
    excel_path = os.path.join(OUTPUT_DIR, excel_filename)

    saved_excel_path = excel_path
    saved_excel_filename = excel_filename
    
    wb = openpyxl.Workbook()
    
    # 폰트 및 스타일 정의
    font_title = Font(name="Malgun Gothic", size=16, bold=True, color="1B365D")
    font_header = Font(name="Malgun Gothic", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Malgun Gothic", size=10, bold=True)
    font_regular = Font(name="Malgun Gothic", size=10)
    
    fill_header = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    fill_summary_hdr = PatternFill(start_color="4A607A", end_color="4A607A", fill_type="solid")
    fill_total = PatternFill(start_color="E9EEF4", end_color="E9EEF4", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    double_bottom_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='double', color='1B365D')
    )

    # ------------------ Tab 1: 종합 집계표 ------------------
    ws1 = wb.active
    ws1.title = "종합 집계표"
    ws1.views.sheetView[0].showGridLines = True
    
    ws1["A1"] = "H빔 기둥 및 보 적산 최종 집계표"
    ws1["A1"].font = font_title
    ws1.row_dimensions[1].height = 30
    
    # 요약 테이블 헤더 (규격 칼럼 추가)
    headers_summary = ["부재 구분", "부재 부호", "H빔 규격 (상세정보)", "수량 (개/개소)", "총 길이 (mm)", "단위 중량 (ton/m)", "총 중량 (ton)"]
    ws1.append([]) # 빈 줄
    ws1.append(headers_summary)
    ws1.row_dimensions[3].height = 25
    for col_idx in range(1, 8):
        cell = ws1.cell(row=3, column=col_idx)
        cell.font = font_header
        cell.fill = fill_summary_hdr
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    # 시트 매핑 사전 구축
    sheets_dict = {s["id"]: s for s in request.sheets}
    
    def get_spec_detail(sheet_id, text):
        sheet = sheets_dict.get(sheet_id, {})
        sched = sheet.get("schedule_table", {}) or {}
        if not text: return "-"
        key = text.upper().strip()
        match = sched.get(key, {})
        if not match:
            for k, v in sched.items():
                if k.replace(" ", "") == key.replace(" ", ""):
                    match = v
                    break
        return match.get("detail", "-").strip()

    def get_hbeam_unit_weight(spec_str, symbol, is_column=True):
        spec = spec_str.upper().replace(" ", "")
        if "300X150" in spec: return 36.7
        if "400X200" in spec: return 66.0
        if "250X250" in spec: return 72.4
        if "440X300" in spec: return 124.0
        if "100X100" in spec: return 17.2
        if "200X200" in spec: return 49.9
        if "350X350" in spec: return 137.0
        if "400X400" in spec: return 172.0
        if "450X200" in spec: return 76.0
        if "500X200" in spec: return 89.6
        if "588X300" in spec: return 151.0
        if "600X200" in spec: return 106.0
        if "700X300" in spec: return 185.0
        if "800X300" in spec: return 210.0
        if "900X300" in spec: return 243.0
        
        # 규격 숫자로 동적 중량 유추
        import re
        nums = re.findall(r'\d+(?:\.\d+)?', spec)
        if len(nums) >= 4:
            try:
                h = float(nums[0])
                w = float(nums[1])
                t1 = float(nums[2])
                t2 = float(nums[3])
                area_mm2 = h * t1 + 2 * w * t2 - 2 * t1 * t2
                weight = (area_mm2 / 100.0) * 7.85 * 0.1
                return round(weight, 1)
            except Exception:
                pass
                
        col_weights = {"C1": 172.0, "C2": 172.0, "C3": 283.0, "SC1": 172.0, "SC2": 283.0, "MC1": 172.0, "MC2": 283.0}
        beam_weights = {"G1": 172.0, "G2": 137.0, "G3": 93.0, "B1": 93.0, "B2": 72.0, "SG1": 137.0, "SG2": 93.0}
        sym = symbol.upper().strip()
        if is_column:
            return col_weights.get(sym, 172.0)
        else:
            return beam_weights.get(sym, 137.0)

    # 부재 구분, 부호, 규격 기반의 통합 집계 딕셔너리
    # key: (부재구분, 부호, 규격) -> value: {count, length, uw}
    takeoff_summary = {}
    
    # 기둥 집계
    for col in request.columns:
        sheet_id = col.get("sheet_id")
        txt = col["text"].upper().strip()
        spec = get_spec_detail(sheet_id, txt)
        h_mm = float(col.get("height", 0))  # mm 단위
        uw = get_hbeam_unit_weight(spec, txt, is_column=True)
        
        key = ("H빔 기둥", txt, spec)
        if key not in takeoff_summary:
            takeoff_summary[key] = {"count": 0, "length_mm": 0.0, "uw": uw}
        takeoff_summary[key]["count"] += 1
        takeoff_summary[key]["length_mm"] += h_mm

    # 보 집계
    # - source 없음: 규칙 기반 보 (중량 계산 포함)
    # - source "both": 규칙 기반 보가 AI와 매칭됨 (중량 계산 포함)
    # - source "ai": AI가 새로 발견한 보 (중량 계산 포함)
    for beam in request.beams:
        sheet_id = beam.get("sheet_id")
        txt = beam["text"].upper().strip()
        spec = get_spec_detail(sheet_id, txt)
        l_mm = float(beam.get("length", 0))  # mm 단위
        uw = get_hbeam_unit_weight(spec, txt, is_column=False)
        
        key = ("H빔 보", txt, spec)
        if key not in takeoff_summary:
            takeoff_summary[key] = {"count": 0, "length_mm": 0.0, "uw": uw}
        takeoff_summary[key]["count"] += 1
        takeoff_summary[key]["length_mm"] += l_mm

    start_row = 4
    curr_row = start_row
    
    total_count = 0
    total_length = 0.0
    total_uw = 0.0
    total_weight = 0.0
    
    # 정렬하여 종합 집계표에 기입 (수식 대신 실제 수치 기입하여 뷰어 호환성 해결)
    for key in sorted(takeoff_summary.keys()):
        kind, mark, spec = key
        data = takeoff_summary[key]
        cnt = data["count"]
        length_mm = data["length_mm"]
        uw = data["uw"]
        weight = (length_mm / 1000.0) * uw / 1000.0  # ton (mm → m → ton)
        
        total_count += cnt
        total_length += length_mm
        total_uw += uw
        total_weight += weight
        
        ws1.append([kind, mark, spec, cnt, round(length_mm, 0), round(uw / 1000.0, 4), round(weight, 3)])
        curr_row += 1
        
    # 서식 적용
    for r in range(start_row, curr_row):
        ws1.row_dimensions[r].height = 20
        for c in range(1, 8):
            cell = ws1.cell(row=r, column=c)
            cell.font = font_regular
            cell.border = thin_border
            if c in [1, 2, 3]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c in [4, 5, 6, 7]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if c == 4:
                    cell.number_format = '#,##0'
                elif c == 5:
                    cell.number_format = '#,##0'
                elif c == 6:
                    cell.number_format = '#,##0.0000'
                elif c == 7:
                    cell.number_format = '#,##0.000'

    # 합계 행 추가 (수량, 총길이, 단위중량, 총중량에 대한 합계 적용)
    ws1.append(["합계", "", "", total_count, round(total_length, 0), round(total_uw / 1000.0, 4), round(total_weight, 3)])
    ws1.row_dimensions[curr_row].height = 22
    for c in range(1, 8):
        cell = ws1.cell(row=curr_row, column=c)
        cell.font = font_bold
        cell.fill = fill_total
        cell.border = double_bottom_border
        if c in [4, 5, 6, 7]:
            cell.alignment = Alignment(horizontal="right", vertical="center")
            if c == 4:
                cell.number_format = '#,##0'
            elif c == 5:
                cell.number_format = '#,##0'
            elif c == 6:
                cell.number_format = '#,##0.0000'
            elif c == 7:
                cell.number_format = '#,##0.000'
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # ------------------ Tab 2: H빔 기둥 산출 근거 ------------------
    ws2 = wb.create_sheet(title="H빔 기둥 산출 근거")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2["A1"] = "H빔 기둥 개별 산출 상세 내역서"
    ws2["A1"].font = font_title
    ws2.row_dimensions[1].height = 30
    
    # H빔 규격 정보 칼럼 추가 (단위 통일: ton)
    headers_col = ["시트명", "부재 부호", "H빔 규격 (상세정보)", "수량 (개)", "설정 높이 (mm)", "단위 중량 (ton/m)", "중량 (ton)", "중심 좌표 X", "중심 좌표 Y"]
    ws2.append([])
    ws2.append(headers_col)
    ws2.row_dimensions[3].height = 25
    for col_idx in range(1, 10):
        cell = ws2.cell(row=3, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    start_row_col = 4
    curr_row_col = start_row_col
    
    sheet_names = {s["id"]: s["name"] for s in request.sheets}
    
    total_col_count = 0
    total_col_weight = 0.0
    
    for idx, col in enumerate(request.columns):
        sheet_id = col["sheet_id"]
        sheet_name = sheet_names.get(sheet_id, "알 수 없음")
        txt = col["text"].upper().strip()
        spec = get_spec_detail(sheet_id, txt)
        h_mm = float(col.get("height", 0))
        uw = get_hbeam_unit_weight(spec, txt, is_column=True)
        weight = (h_mm / 1000.0) * uw / 1000.0  # ton 단위 (종합집계표와 통일)
        weight_rounded = round(weight, 3)  # 반올림된 값 (엑셀 표시와 동일)
        
        total_col_count += 1
        total_col_weight += weight_rounded  # 반올림된 값 누적 (표시값 합계와 일치)
        
        ws2.append([
            sheet_name,
            txt,
            spec,
            1,
            h_mm,
            round(uw / 1000.0, 4),
            weight_rounded,
            round(col["cx"], 1),
            round(col["cy"], 1)
        ])
        curr_row_col += 1
        
    for r in range(start_row_col, curr_row_col):
        ws2.row_dimensions[r].height = 20
        for c in range(1, 10):
            cell = ws2.cell(row=r, column=c)
            cell.font = font_regular
            cell.border = thin_border
            if c in [1, 2, 3]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c in [4, 5, 6, 7, 8, 9]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if c in [4, 5]:
                    cell.number_format = '#,##0'
                elif c == 6:
                    cell.number_format = '#,##0.0000'
                elif c == 7:
                    cell.number_format = '#,##0.000'
                elif c in [8, 9]:
                    cell.number_format = '#,##0'
 
    # 기둥 합계 행 (ton 단위, 종합집계표와 통일)
    if curr_row_col > start_row_col:
        ws2.append(["합계", "", "", total_col_count, "", "", round(total_col_weight, 3), "", ""])
        ws2.row_dimensions[curr_row_col].height = 22
        for c in range(1, 10):
            cell = ws2.cell(row=curr_row_col, column=c)
            cell.font = font_bold
            cell.fill = fill_total
            cell.border = double_bottom_border
            if c in [4, 7]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if c == 4:
                    cell.number_format = '#,##0'
                else:
                    cell.number_format = '#,##0.000'
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # ------------------ Tab 3: H빔 보 산출 근거 ------------------
    ws3 = wb.create_sheet(title="H빔 보 산출 근거")
    ws3.views.sheetView[0].showGridLines = True
    
    ws3["A1"] = "H빔 보 개별 산출 상세 내역서"
    ws3["A1"].font = font_title
    ws3.row_dimensions[1].height = 30
    
    # H빔 규격 정보 칼럼 추가 (길이(m) 제거 및 중량 계산 조정)
    headers_beam = ["시트명", "부재 부호", "H빔 규격 (상세정보)", "실측 길이 (mm)", "단위 중량 (ton/m)", "중량 (ton)", "시작점 X", "시작점 Y", "끝점 X", "끝점 Y"]
    ws3.append([])
    ws3.append(headers_beam)
    ws3.row_dimensions[3].height = 25
    for col_idx in range(1, 11):
        cell = ws3.cell(row=3, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    start_row_beam = 4
    curr_row_beam = start_row_beam
    
    total_beam_length = 0.0
    total_beam_weight = 0.0
    
    for idx, beam in enumerate(request.beams):
        sheet_id = beam["sheet_id"]
        sheet_name = sheet_names.get(sheet_id, "알 수 없음")
        txt = beam["text"].upper().strip()
        spec = get_spec_detail(sheet_id, txt)
        l_mm = float(beam.get("length", 0))
        uw = get_hbeam_unit_weight(spec, txt, is_column=False)
        weight = (l_mm / 1000.0) * uw / 1000.0  # ton 단위 (종합집계표와 통일)
        weight_rounded = round(weight, 3)  # 반올림된 값 (엑셀 표시와 동일)
        
        total_beam_length += l_mm
        total_beam_weight += weight_rounded  # 반올림된 값 누적 (표시값 합계와 일치)
        
        ws3.append([
            sheet_name,
            txt,
            spec,
            l_mm,
            round(uw / 1000.0, 4),
            weight_rounded,
            round(beam["start"][0], 1),
            round(beam["start"][1], 1),
            round(beam["end"][0], 1),
            round(beam["end"][1], 1)
        ])
        curr_row_beam += 1
        
    for r in range(start_row_beam, curr_row_beam):
        ws3.row_dimensions[r].height = 20
        for c in range(1, 11):
            cell = ws3.cell(row=r, column=c)
            cell.font = font_regular
            cell.border = thin_border
            if c in [1, 2, 3]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c in [4, 5, 6, 7, 8, 9, 10]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if c in [4, 7, 8, 9, 10]:
                    cell.number_format = '#,##0'
                elif c == 5:
                    cell.number_format = '#,##0.0000'
                elif c == 6:
                    cell.number_format = '#,##0.000'

    # 보 합계 행 (길이(m) 제거로 열 개수 10개에 대응, 실측길이와 중량 합계 반영)
    if curr_row_beam > start_row_beam:
        ws3.append(["합계", "", "", round(total_beam_length, 0), "", round(total_beam_weight, 3), "", "", "", ""])
        ws3.row_dimensions[curr_row_beam].height = 22
        for c in range(1, 11):
            cell = ws3.cell(row=curr_row_beam, column=c)
            cell.font = font_bold
            cell.fill = fill_total
            cell.border = double_bottom_border
            if c in [4, 6]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if c == 4:
                    cell.number_format = '#,##0'
                else:
                    cell.number_format = '#,##0.000'
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # 엑셀 쓰기 충돌 방지 로직 (최대 10회 우회 시도)
    saved = False
    for attempt in range(10):
        try:
            wb.save(saved_excel_path)
            saved = True
            break
        except PermissionError:
            saved_excel_filename = request.filename.replace(".dxf", f"_takeoff_new_{attempt + 1}.xlsx")
            saved_excel_path = os.path.join(OUTPUT_DIR, saved_excel_filename)
            
    if not saved:
        raise HTTPException(status_code=500, detail="PermissionError: Failed to save Excel. Excel file might be opened by another program.")

    # 3. PDF 산출 근거 보고서 생성
    pdf_filename = request.filename.replace(".dxf", "_takeoff.pdf")
    pdf_path = os.path.join(OUTPUT_DIR, pdf_filename)
    
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.graphics.shapes import Drawing, Line as PDFLine, Circle as PDFCircle, String as PDFString, Rect as PDFRect
        
        # 맑은 고딕 폰트 등록
        font_registered = False
        bold_font_registered = False
        system_root = os.environ.get('SystemRoot', 'C:\\Windows')
        
        font_file = os.path.join(system_root, 'Fonts', 'malgun.ttf')
        if not os.path.exists(font_file):
            font_file = 'C:\\Windows\\Fonts\\malgun.ttf'
            
        bold_font_file = os.path.join(system_root, 'Fonts', 'malgunbd.ttf')
        if not os.path.exists(bold_font_file):
            bold_font_file = 'C:\\Windows\\Fonts\\malgunbd.ttf'
            
        if os.path.exists(font_file):
            pdfmetrics.registerFont(TTFont('MalgunGothic', font_file))
            font_registered = True
            
        if os.path.exists(bold_font_file):
            pdfmetrics.registerFont(TTFont('MalgunGothic-Bold', bold_font_file))
            bold_font_registered = True
        
        font_name = 'MalgunGothic' if font_registered else 'Helvetica'
        bold_font_name = 'MalgunGothic-Bold' if bold_font_registered else ('Helvetica-Bold' if font_name == 'Helvetica' else font_name)
        
        doc = SimpleDocTemplate(
            pdf_path,
            pagesize=A4,
            leftMargin=36,
            rightMargin=36,
            topMargin=36,
            bottomMargin=36
        )
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'PDFTitle',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=16,
            leading=20,
            textColor=colors.HexColor('#1B365D'),
            spaceAfter=15,
            alignment=1
        )
        subtitle_style = ParagraphStyle(
            'PDFSubtitle',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=12,
            leading=16,
            textColor=colors.HexColor('#4A607A'),
            spaceAfter=10,
            alignment=0
        )
        normal_style = ParagraphStyle(
            'PDFNormal',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=9,
            leading=12,
            textColor=colors.HexColor('#333333')
        )
        header_style = ParagraphStyle(
            'PDFHeader',
            parent=styles['Normal'],
            fontName=font_name,
            fontSize=9,
            leading=12,
            textColor=colors.white,
            alignment=1
        )
        
        story = []
        
        # 표지 및 종합 집계표 추가
        story.append(Paragraph("H빔 적산 산출 근거 보고서", title_style))
        story.append(Paragraph(f"원 도면 파일명: {request.filename}", subtitle_style))
        story.append(Spacer(1, 10))
        
        summary_table_data = [[
            Paragraph("<b>부재 구분</b>", header_style),
            Paragraph("<b>부재 부호</b>", header_style),
            Paragraph("<b>H빔 규격 (상세정보)</b>", header_style),
            Paragraph("<b>수량 (개)</b>", header_style),
            Paragraph("<b>총 길이 (mm)</b>", header_style),
            Paragraph("<b>단위중량 (ton/m)</b>", header_style),
            Paragraph("<b>총 중량 (ton)</b>", header_style)
        ]]
        
        total_count = 0
        total_length = 0.0
        total_weight = 0.0
        
        for key in sorted(takeoff_summary.keys()):
            kind, mark, spec = key
            data = takeoff_summary[key]
            cnt = data["count"]
            length_mm = data["length_mm"]
            uw = data["uw"]
            weight = (length_mm / 1000.0) * uw / 1000.0  # ton (mm → m → ton)
            
            total_count += cnt
            total_length += length_mm
            total_weight += weight
            
            summary_table_data.append([
                Paragraph(kind, normal_style),
                Paragraph(mark, normal_style),
                Paragraph(spec, normal_style),
                Paragraph(f"{cnt:,}", normal_style),
                Paragraph(f"{length_mm:,.0f}", normal_style),
                Paragraph(f"{uw / 1000.0:,.4f}", normal_style),
                Paragraph(f"{weight:,.3f}", normal_style)
            ])
            
        summary_table_data.append([
            Paragraph("<b>합계</b>", normal_style),
            Paragraph("", normal_style),
            Paragraph("", normal_style),
            Paragraph(f"<b>{total_count:,}</b>", normal_style),
            Paragraph(f"<b>{total_length:,.0f}</b>", normal_style),
            Paragraph("", normal_style),
            Paragraph(f"<b>{total_weight:,.3f}</b>", normal_style)
        ])
        
        t_summary = Table(summary_table_data, colWidths=[70, 60, 150, 50, 60, 63, 70])
        t_summary.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1B365D')),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D9D9D9')),
            ('ROWBACKGROUNDS', (0,1), (-1,-2), [colors.white, colors.HexColor('#F8FAFC')]),
            ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#E9EEF4')),
            ('LINEBELOW', (0,-1), (-1,-1), 1.5, colors.HexColor('#1B365D')),
        ]))
        
        story.append(t_summary)
        story.append(PageBreak())
        
        # 각 승인 시트별 도면 및 개별 자재 명세 렌더링
        for sheet in request.sheets:
            sheet_id = sheet["id"]
            sheet_name = sheet["name"]
            sheet_number = sheet.get("number", "")
            
            story.append(Paragraph(f"산출 근거 도면 - [{sheet_number}] {sheet_name}", title_style))
            
            bbox = sheet.get("bbox")
            has_main_drawing = False
            if bbox and len(bbox) == 4:
                xmin, ymin, xmax, ymax = bbox
                sw = xmax - xmin
                sh = ymax - ymin
                
                draw_w = 523
                draw_h = 350
                
                if sw > 0 and sh > 0:
                    has_main_drawing = True
                    scale = min(draw_w / sw, draw_h / sh)
                    
                    d = Drawing(draw_w, draw_h)
                    d.add(PDFRect(0, 0, draw_w, draw_h, fillColor=colors.HexColor('#020617'), strokeColor=colors.HexColor('#1E293B')))
                    
                    def tx(wx): return (wx - xmin) * scale
                    def ty(wy): return (wy - ymin) * scale
                    
                    t_lines = sheet.get("thumbnail_lines", [])
                    for ln in t_lines:
                        is_green = ln.get("is_green_guide")
                        stroke_color = colors.Color(16/255.0, 185/255.0, 129/255.0, 0.5) if is_green else colors.Color(91/255.0, 192/255.0, 190/255.0, 0.3)
                        stroke_width = 0.5 if is_green else 0.4
                        
                        start_pt = ln.get("start")
                        end_pt = ln.get("end")
                        if start_pt and end_pt:
                            x1, y1 = start_pt
                            x2, y2 = end_pt
                            d.add(PDFLine(tx(x1), ty(y1), tx(x2), ty(y2), strokeColor=stroke_color, strokeWidth=stroke_width))
                            
                    sheet_cols = [c for c in request.columns if c["sheet_id"] == sheet_id]
                    for col in sheet_cols:
                        cx, cy = col["cx"], col["cy"]
                        txt = col.get("text", "")
                        col_source = col.get("source", "")
                        is_ai = col_source == "ai"
                        is_both = col_source == "both"
                        
                        if is_both:
                            # Both 기둥: 룰베이스(속이 찬 원) + AI(속이 빈 원)
                            d.add(PDFCircle(tx(cx), ty(cy), 3.5, fillColor=colors.HexColor('#F59E0B'), strokeColor=colors.white, strokeWidth=0.3))
                            # AI 점선 원 (더 크게)
                            ai_circle_r = 7.0
                            ai_cx_pdf = tx(cx)
                            ai_cy_pdf = ty(cy)
                            num_segments = 12
                            for i in range(num_segments):
                                angle1 = 2 * math.pi * i / num_segments
                                angle2 = 2 * math.pi * (i + 0.6) / num_segments
                                x1_seg = ai_cx_pdf + ai_circle_r * math.cos(angle1)
                                y1_seg = ai_cy_pdf + ai_circle_r * math.sin(angle1)
                                x2_seg = ai_cx_pdf + ai_circle_r * math.cos(angle2)
                                y2_seg = ai_cy_pdf + ai_circle_r * math.sin(angle2)
                                d.add(PDFLine(x1_seg, y1_seg, x2_seg, y2_seg, strokeColor=colors.HexColor('#FF007F'), strokeWidth=1.0))
                            d.add(PDFString(tx(cx) + 4, ty(cy) - 3.5, txt, fontName=bold_font_name, fontSize=8, fillColor=colors.HexColor('#FF007F')))
                        elif is_ai:
                            # AI 기둥: 속이 빈 핫핑크 점선 원 (더 크게)
                            ai_circle_r = 7.0
                            ai_cx_pdf = tx(cx)
                            ai_cy_pdf = ty(cy)
                            num_segments = 12
                            for i in range(num_segments):
                                angle1 = 2 * math.pi * i / num_segments
                                angle2 = 2 * math.pi * (i + 0.6) / num_segments
                                x1_seg = ai_cx_pdf + ai_circle_r * math.cos(angle1)
                                y1_seg = ai_cy_pdf + ai_circle_r * math.sin(angle1)
                                x2_seg = ai_cx_pdf + ai_circle_r * math.cos(angle2)
                                y2_seg = ai_cy_pdf + ai_circle_r * math.sin(angle2)
                                d.add(PDFLine(x1_seg, y1_seg, x2_seg, y2_seg, strokeColor=colors.HexColor('#FF007F'), strokeWidth=1.0))
                            d.add(PDFString(tx(cx) + 4, ty(cy) - 3.5, txt, fontName=bold_font_name, fontSize=8, fillColor=colors.HexColor('#FF007F')))
                        else:
                            # 룰베이스 기둥: 속이 찬 원
                            d.add(PDFCircle(tx(cx), ty(cy), 3.5, fillColor=colors.HexColor('#F59E0B'), strokeColor=colors.white, strokeWidth=0.3))
                            d.add(PDFString(tx(cx) + 4, ty(cy) - 3.5, txt, fontName=bold_font_name, fontSize=8, fillColor=colors.HexColor('#F59E0B')))
                        
                    sheet_beams = [b for b in request.beams if b["sheet_id"] == sheet_id]
                    
                    # zoo guide: PDF 좌표 기준으로 보 그리기
                    # 디버그: AI 보 확인
                    ai_beams_in_sheet = [b for b in sheet_beams if b.get("source") == "ai"]
                    both_beams_in_sheet = [b for b in sheet_beams if b.get("source") == "both"]
                    print(f"[PDF] 시트 {sheet_id}: 전체 보 {len(sheet_beams)}개, AI 보 {len(ai_beams_in_sheet)}개, Both 보 {len(both_beams_in_sheet)}개")

                    for beam in sheet_beams:
                        start_pt = beam.get("start")
                        end_pt = beam.get("end")
                        txt = beam.get("text", "")
                        if start_pt and end_pt:
                            x1, y1 = start_pt
                            x2, y2 = end_pt
                            beam_source = beam.get("source", "")
                            is_ai = beam_source == "ai"
                            is_both = beam_source == "both"
                            
                            if is_both:
                                # Both 보: 룰베이스(하늘색 실선+속이찬 원) + AI(핫핑크 틱선+속이빈 원)
                                # 1. 하늘색 실선 (룰베이스)
                                d.add(PDFLine(tx(x1), ty(y1), tx(x2), ty(y2), strokeColor=colors.HexColor('#06B6D4'), strokeWidth=1.0))
                                
                                # 2. 핫핑크 틱선 (AI)
                                dx = x2 - x1
                                dy = y2 - y1
                                seg_len = math.sqrt(dx**2 + dy**2)
                                if seg_len > 0:
                                    tick_w_pdf = 5.0
                                    start_pdf_x = tx(x1)
                                    start_pdf_y = ty(y1)
                                    end_pdf_x = tx(x2)
                                    end_pdf_y = ty(y2)
                                    pdf_dx = end_pdf_x - start_pdf_x
                                    pdf_dy = end_pdf_y - start_pdf_y
                                    pdf_seg_len = math.sqrt(pdf_dx**2 + pdf_dy**2)
                                    
                                    if pdf_seg_len > 0:
                                        pdf_nx = -pdf_dy / pdf_seg_len
                                        pdf_ny = pdf_dx / pdf_seg_len
                                        
                                        # 시작 틱 (AI, 얇게: 1.0 → 0.7)
                                        d.add(PDFLine(start_pdf_x + pdf_nx * tick_w_pdf, start_pdf_y + pdf_ny * tick_w_pdf,
                                                    start_pdf_x - pdf_nx * tick_w_pdf, start_pdf_y - pdf_ny * tick_w_pdf,
                                                    strokeColor=colors.HexColor('#FF007F'), strokeWidth=0.7))
                                        # 끝 틱 (AI, 얇게: 1.0 → 0.7)
                                        d.add(PDFLine(end_pdf_x + pdf_nx * tick_w_pdf, end_pdf_y + pdf_ny * tick_w_pdf,
                                                    end_pdf_x - pdf_nx * tick_w_pdf, end_pdf_y - pdf_ny * tick_w_pdf,
                                                    strokeColor=colors.HexColor('#FF007F'), strokeWidth=0.7))
                                
                                # 3. 중앙 원형 포인트: 룰베이스(속이 찬 하늘색 원) + AI(속이 빈 핫핑크 점선 원)
                                mx, my = (x1 + x2) / 2.0, (y1 + y2) / 2.0
                                # 룰베이스: 속이 찬 하늘색 원
                                d.add(PDFCircle(tx(mx), ty(my), 3.5, fillColor=colors.HexColor('#06B6D4'), strokeColor=colors.white, strokeWidth=0.3))
                                # AI: 속이 빈 핫핑크 점선 원 (기둥과 동일하게 크게)
                                ai_circle_r = 7.0
                                ai_cx_pdf = tx(mx)
                                ai_cy_pdf = ty(my)
                                num_segments = 12
                                for i in range(num_segments):
                                    angle1 = 2 * math.pi * i / num_segments
                                    angle2 = 2 * math.pi * (i + 0.6) / num_segments
                                    x1_seg = ai_cx_pdf + ai_circle_r * math.cos(angle1)
                                    y1_seg = ai_cy_pdf + ai_circle_r * math.sin(angle1)
                                    x2_seg = ai_cx_pdf + ai_circle_r * math.cos(angle2)
                                    y2_seg = ai_cy_pdf + ai_circle_r * math.sin(angle2)
                                    d.add(PDFLine(x1_seg, y1_seg, x2_seg, y2_seg, strokeColor=colors.HexColor('#FF007F'), strokeWidth=1.0))
                                d.add(PDFString(tx(mx), ty(my) + 4, txt, fontName=bold_font_name, fontSize=8, fillColor=colors.HexColor('#FF007F')))
                                
                            elif is_ai:
                                # AI 승인 보: 전체 선 대신 양 끝단 핫핑크 직교 틱선 + 중앙 핫핑크 원형 포인트
                                dx = x2 - x1
                                dy = y2 - y1
                                seg_len = math.sqrt(dx**2 + dy**2)
                                if seg_len > 0:
                                    nx = -dy / seg_len
                                    ny = dx / seg_len
                                    # tick_w = 150.0  # 월드 좌표 기준 150mm 반폭 틱
                                    
                                    # # 시작 틱
                                    # d.add(PDFLine(tx(x1 + nx*tick_w), ty(y1 + ny*tick_w), tx(x1 - nx*tick_w), ty(y1 - ny*tick_w), strokeColor=colors.HexColor('#FF007F'), strokeWidth=1.2))
                                    # # 끝 틱
                                    # d.add(PDFLine(tx(x2 + nx*tick_w), ty(y2 + ny*tick_w), tx(x2 - nx*tick_w), ty(y2 - ny*tick_w), strokeColor=colors.HexColor('#FF007F'), strokeWidth=1.2))
                                    
                                    # zoo guide: PDF 좌표 기준으로 틱 크기 조정
                                    # PDF 좌표로 변환 후, PDF 좌표 기준으로 8mm �반폭 틱 생성
                                    # PDF 좌표 기준 틱 크기 (8mm)
                                    tick_w_pdf = 5.0

                                    # PDF 좌표로 변환
                                    start_pdf_x = tx(x1)
                                    start_pdf_y = ty(y1)
                                    end_pdf_x = tx(x2)
                                    end_pdf_y = ty(y2)

                                    # PDF 좌표 기준 방향 벡터
                                    pdf_dx = end_pdf_x - start_pdf_x
                                    pdf_dy = end_pdf_y - start_pdf_y
                                    pdf_seg_len = math.sqrt(pdf_dx**2 + pdf_dy**2)

                                    if pdf_seg_len > 0:
                                        pdf_nx = -pdf_dy / pdf_seg_len
                                        pdf_ny = pdf_dx / pdf_seg_len
                                        
                                        # 시작 틱 (AI, 얇게: 1.0 → 0.7)
                                        d.add(PDFLine(start_pdf_x + pdf_nx * tick_w_pdf, start_pdf_y + pdf_ny * tick_w_pdf,
                                                    start_pdf_x - pdf_nx * tick_w_pdf, start_pdf_y - pdf_ny * tick_w_pdf,
                                                    strokeColor=colors.HexColor('#FF007F'), strokeWidth=0.7))
                                        # 끝 틱 (AI, 얇게: 1.0 → 0.7)
                                        d.add(PDFLine(end_pdf_x + pdf_nx * tick_w_pdf, end_pdf_y + pdf_ny * tick_w_pdf,
                                                    end_pdf_x - pdf_nx * tick_w_pdf, end_pdf_y - pdf_ny * tick_w_pdf,
                                                    strokeColor=colors.HexColor('#FF007F'), strokeWidth=0.7))

                                
                                # 중앙 원형 포인트 (속이 빈 핫핑크 점선 원, 기둥과 동일하게 크게)
                                mx, my = (x1 + x2) / 2.0, (y1 + y2) / 2.0
                                ai_circle_r = 7.0
                                ai_cx_pdf = tx(mx)
                                ai_cy_pdf = ty(my)
                                num_segments = 12
                                for i in range(num_segments):
                                    angle1 = 2 * math.pi * i / num_segments
                                    angle2 = 2 * math.pi * (i + 0.6) / num_segments
                                    x1_seg = ai_cx_pdf + ai_circle_r * math.cos(angle1)
                                    y1_seg = ai_cy_pdf + ai_circle_r * math.sin(angle1)
                                    x2_seg = ai_cx_pdf + ai_circle_r * math.cos(angle2)
                                    y2_seg = ai_cy_pdf + ai_circle_r * math.sin(angle2)
                                    d.add(PDFLine(x1_seg, y1_seg, x2_seg, y2_seg, strokeColor=colors.HexColor('#FF007F'), strokeWidth=1.0))
                                d.add(PDFString(tx(mx), ty(my) + 4, txt, fontName=bold_font_name, fontSize=8, fillColor=colors.HexColor('#FF007F')))
                            else:
                                # 일반 보(룰베이스): 하늘색 실선 + 속이 찬 하늘색 중앙 원
                                d.add(PDFLine(tx(x1), ty(y1), tx(x2), ty(y2), strokeColor=colors.HexColor('#06B6D4'), strokeWidth=1.0))
                                mx, my = (x1 + x2) / 2.0, (y1 + y2) / 2.0
                                d.add(PDFCircle(tx(mx), ty(my), 3.5, fillColor=colors.HexColor('#06B6D4'), strokeColor=colors.white, strokeWidth=0.3))
                                d.add(PDFString(tx(mx), ty(my) + 4, txt, fontName=bold_font_name, fontSize=8, fillColor=colors.HexColor('#06B6D4')))
                            
                    # 도면 텍스트 및 선택한 수치 강조 그리기
                    t_texts = sheet.get("thumbnail_texts", []) or []
                    
                    # 현재 시트 기둥/보의 높이/길이 값을 AI/일반으로 분리하여 집계
                    normal_h_vals = set()
                    ai_h_vals = set()
                    
                    # 현재 시트 기둥 height 집계 (우선순위)
                    for col in sheet_cols:
                        h_val = col.get("height")
                        if h_val:
                            try:
                                val = float(h_val)
                                if val > 0:
                                    if col.get("source") == "ai":
                                        ai_h_vals.add(val)
                                    else:
                                        normal_h_vals.add(val)
                            except (ValueError, TypeError):
                                pass
                    
                    # 현재 시트 보 length 집계
                    for beam in sheet_beams:
                        l_val = beam.get("length")
                        if l_val:
                            try:
                                val = float(l_val)
                                if val > 0:
                                    if beam.get("source") == "ai":
                                        ai_h_vals.add(val)
                                    else:
                                        normal_h_vals.add(val)
                            except (ValueError, TypeError):
                                pass
                    
                    # 다른 시트 값도 포함 (현재 시트에 없는 경우 fallback)
                    for col in request.columns:
                        if col.get("sheet_id") == sheet_id:
                            continue  # 이미 위에서 처리됨
                        h_val = col.get("height")
                        if h_val:
                            try:
                                val = float(h_val)
                                if val > 0:
                                    if col.get("source") == "ai":
                                        ai_h_vals.add(val)
                                    else:
                                        normal_h_vals.add(val)
                            except (ValueError, TypeError):
                                pass
                    
                    import re
                    for t in t_texts:
                        tx_val = float(t.get("x", 0))
                        ty_val = float(t.get("y", 0))
                        txt_str = t.get("text", "")
                        
                        cx = tx(tx_val)
                        cy = ty(ty_val)
                        
                        if cx < -30 or cx > draw_w + 30 or cy < -30 or cy > draw_h + 30:
                            continue
                        
                        if not txt_str:
                            continue
                            
                        clean = txt_str.strip()
                        # 쉼표 제거 후 숫자 추출 (정수 1000~20000 범위만 유효 높이/길이로 판단)
                        clean_no_comma = clean.replace(",", "")
                        nums_found = re.findall(r'\b(\d{4,6})\b', clean_no_comma)
                        
                        is_ai_val = False
                        is_normal_val = False
                        for num_str in nums_found:
                            try:
                                val = float(num_str)
                                if val in ai_h_vals:
                                    is_ai_val = True
                                    break
                                elif val in normal_h_vals:
                                    is_normal_val = True
                            except (ValueError, TypeError):
                                pass
                                
                        if is_ai_val:
                            # AI 관련 높이/수치는 핫핑크색으로 강조
                            d.add(PDFString(cx, cy, txt_str, fontName=bold_font_name, fontSize=8, fillColor=colors.HexColor('#FF007F')))
                        elif is_normal_val:
                            d.add(PDFString(cx, cy, txt_str, fontName=bold_font_name, fontSize=8, fillColor=colors.HexColor('#F59E0B')))
                        else:
                            d.add(PDFString(cx, cy, txt_str, fontName=font_name, fontSize=5, fillColor=colors.Color(148/255.0, 163/255.0, 184/255.0, 0.6)))
                            
                    story.append(d)
                    story.append(Spacer(1, 10))
            
            # 기둥 높이 산출 근거 도면 추가
            has_ref_drawing = False
            if sheet.get("job") == "column" and sheet.get("height_ref_sheet_id"):
                ref_id = sheet.get("height_ref_sheet_id")
                ref_sheet = sheets_dict.get(ref_id) or all_sheets_dict.get(ref_id)
                if ref_sheet:
                    ref_bbox = ref_sheet.get("bbox")
                    if ref_bbox and len(ref_bbox) == 4:
                        ref_xmin, ref_ymin, ref_xmax, ref_ymax = ref_bbox
                        ref_sw = ref_xmax - ref_xmin
                        ref_sh = ref_ymax - ref_ymin
                        
                        if ref_sw > 0 and ref_sh > 0:
                            has_ref_drawing = True
                            if has_main_drawing:
                                story.append(PageBreak())
                                
                            ref_number = ref_sheet.get("number", "")
                            ref_name = ref_sheet.get("name", "")
                            story.append(Paragraph(f"기둥 높이 산출 근거 도면 - [{ref_number}] {ref_name}", subtitle_style))
                            story.append(Spacer(1, 5))
                            
                            ref_scale = min(draw_w / ref_sw, draw_h / ref_sh)
                            d_ref = Drawing(draw_w, draw_h)
                            d_ref.add(PDFRect(0, 0, draw_w, draw_h, fillColor=colors.HexColor('#020617'), strokeColor=colors.HexColor('#1E293B')))
                            
                            def tx_ref(wx): return (wx - ref_xmin) * ref_scale
                            def ty_ref(wy): return (wy - ref_ymin) * ref_scale
                            
                            ref_lines = ref_sheet.get("thumbnail_lines", [])
                            for ln in ref_lines:
                                is_green = ln.get("is_green_guide")
                                stroke_color = colors.Color(16/255.0, 185/255.0, 129/255.0, 0.5) if is_green else colors.Color(91/255.0, 192/255.0, 190/255.0, 0.3)
                                stroke_width = 0.5 if is_green else 0.4
                                
                                start_pt = ln.get("start")
                                end_pt = ln.get("end")
                                if start_pt and end_pt:
                                    x1, y1 = start_pt
                                    x2, y2 = end_pt
                                    d_ref.add(PDFLine(tx_ref(x1), ty_ref(y1), tx_ref(x2), ty_ref(y2), strokeColor=stroke_color, strokeWidth=stroke_width))
                                    
                            ref_cols = [c for c in request.columns if c["sheet_id"] == ref_id]
                            for col in ref_cols:
                                cx, cy = col["cx"], col["cy"]
                                txt = col.get("text", "")
                                col_source = col.get("source", "")
                                is_ai = col_source == "ai"
                                is_both = col_source == "both"
                                
                                if is_both:
                                    d_ref.add(PDFCircle(tx_ref(cx), ty_ref(cy), 3.5, fillColor=colors.HexColor('#F59E0B'), strokeColor=colors.white, strokeWidth=0.3))
                                    # AI 점선 원 (더 크게)
                                    ai_circle_r = 7.0
                                    ai_cx_pdf = tx_ref(cx)
                                    ai_cy_pdf = ty_ref(cy)
                                    num_segments = 12
                                    for i in range(num_segments):
                                        angle1 = 2 * math.pi * i / num_segments
                                        angle2 = 2 * math.pi * (i + 0.6) / num_segments
                                        x1_seg = ai_cx_pdf + ai_circle_r * math.cos(angle1)
                                        y1_seg = ai_cy_pdf + ai_circle_r * math.sin(angle1)
                                        x2_seg = ai_cx_pdf + ai_circle_r * math.cos(angle2)
                                        y2_seg = ai_cy_pdf + ai_circle_r * math.sin(angle2)
                                        d_ref.add(PDFLine(x1_seg, y1_seg, x2_seg, y2_seg, strokeColor=colors.HexColor('#FF007F'), strokeWidth=1.0))
                                    d_ref.add(PDFString(tx_ref(cx) + 4, ty_ref(cy) - 3.5, txt, fontName=bold_font_name, fontSize=8, fillColor=colors.HexColor('#FF007F')))
                                elif is_ai:
                                    # AI 기둥: 속이 빈 핫핑크 점선 원 (더 크게)
                                    ai_circle_r = 7.0
                                    ai_cx_pdf = tx_ref(cx)
                                    ai_cy_pdf = ty_ref(cy)
                                    num_segments = 12
                                    for i in range(num_segments):
                                        angle1 = 2 * math.pi * i / num_segments
                                        angle2 = 2 * math.pi * (i + 0.6) / num_segments
                                        x1_seg = ai_cx_pdf + ai_circle_r * math.cos(angle1)
                                        y1_seg = ai_cy_pdf + ai_circle_r * math.sin(angle1)
                                        x2_seg = ai_cx_pdf + ai_circle_r * math.cos(angle2)
                                        y2_seg = ai_cy_pdf + ai_circle_r * math.sin(angle2)
                                        d_ref.add(PDFLine(x1_seg, y1_seg, x2_seg, y2_seg, strokeColor=colors.HexColor('#FF007F'), strokeWidth=1.0))
                                    d_ref.add(PDFString(tx_ref(cx) + 4, ty_ref(cy) - 3.5, txt, fontName=bold_font_name, fontSize=8, fillColor=colors.HexColor('#FF007F')))
                                else:
                                    d_ref.add(PDFCircle(tx_ref(cx), ty_ref(cy), 3.5, fillColor=colors.HexColor('#F59E0B'), strokeColor=colors.white, strokeWidth=0.3))
                                    d_ref.add(PDFString(tx_ref(cx) + 4, ty_ref(cy) - 3.5, txt, fontName=bold_font_name, fontSize=8, fillColor=colors.HexColor('#F59E0B')))
                                
                            # 높이 텍스트 강조 그리기 (참조 도면)
                            ref_texts = ref_sheet.get("thumbnail_texts", []) or []
                            sheet_cols_for_h = [c for c in request.columns if c["sheet_id"] == sheet_id]
                            
                            ref_normal_h_vals = set()
                            ref_ai_h_vals = set()
                            for c in sheet_cols_for_h:
                                h_val = c.get("height")
                                if h_val:
                                    try:
                                        val = float(h_val)
                                        if val > 0:
                                            col_source = c.get("source", "")
                                            if col_source == "ai" or col_source == "both":
                                                ref_ai_h_vals.add(val)
                                            else:
                                                ref_normal_h_vals.add(val)
                                    except (ValueError, TypeError):
                                        pass
                            
                            import re
                            for t in ref_texts:
                                tx_val = float(t.get("x", 0))
                                ty_val = float(t.get("y", 0))
                                txt_str = t.get("text", "")
                                
                                if not txt_str:
                                    continue
                                
                                cx = tx_ref(tx_val)
                                cy = ty_ref(ty_val)
                                
                                if cx < -30 or cx > draw_w + 30 or cy < -30 or cy > draw_h + 30:
                                    continue
                                    
                                clean = txt_str.strip()
                                clean_no_comma = clean.replace(",", "")
                                nums_found = re.findall(r'\b(\d{4,6})\b', clean_no_comma)
                                is_ai_val = False
                                is_normal_val = False
                                for num_str in nums_found:
                                    try:
                                        val = float(num_str)
                                        if val in ref_ai_h_vals:
                                            is_ai_val = True
                                            break
                                        elif val in ref_normal_h_vals:
                                            is_normal_val = True
                                    except (ValueError, TypeError):
                                        pass
                                        
                                if is_ai_val:
                                    # AI 높이 텍스트: 핫핑크 진하게 + 글씨를 포함하는 점선 동그라미
                                    # 텍스트 길이에 따라 동그라미 크기 조정
                                    text_len = len(txt_str)
                                    circle_r = max(10.0, text_len * 2.5 + 5.0)
                                    # 점선 원 그리기
                                    num_segments = 12
                                    for i in range(num_segments):
                                        angle1 = 2 * math.pi * i / num_segments
                                        angle2 = 2 * math.pi * (i + 0.6) / num_segments
                                        x1_seg = cx + circle_r * math.cos(angle1)
                                        y1_seg = cy + circle_r * math.sin(angle1)
                                        x2_seg = cx + circle_r * math.cos(angle2)
                                        y2_seg = cy + circle_r * math.sin(angle2)
                                        d_ref.add(PDFLine(x1_seg, y1_seg, x2_seg, y2_seg, strokeColor=colors.HexColor('#FF007F'), strokeWidth=1.0))
                                    d_ref.add(PDFString(cx, cy, txt_str, fontName=bold_font_name, fontSize=10, fillColor=colors.HexColor('#FF007F')))
                                elif is_normal_val:
                                    # 룰베이스 높이 텍스트: 주황색 진하게
                                    d_ref.add(PDFString(cx, cy, txt_str, fontName=bold_font_name, fontSize=10, fillColor=colors.HexColor('#F59E0B')))
                                else:
                                    d_ref.add(PDFString(cx, cy, txt_str, fontName=font_name, fontSize=5, fillColor=colors.Color(148/255.0, 163/255.0, 184/255.0, 0.6)))
                                    
                            story.append(d_ref)
                            story.append(Spacer(1, 15))
            
            if has_ref_drawing:
                story.append(PageBreak())
                
            sheet_cols = [c for c in request.columns if c["sheet_id"] == sheet_id]
            sheet_beams = [b for b in request.beams if b["sheet_id"] == sheet_id]
            
            details_table_data = [[
                Paragraph("<b>부재 분류</b>", header_style),
                Paragraph("<b>부재 부호</b>", header_style),
                Paragraph("<b>H빔 규격 (상세정보)</b>", header_style),
                Paragraph("<b>길이/높이 (mm)</b>", header_style),
                Paragraph("<b>단위중량 (ton/m)</b>", header_style),
                Paragraph("<b>총 중량 (ton)</b>", header_style),
                Paragraph("<b>X 좌표</b>", header_style),
                Paragraph("<b>Y 좌표</b>", header_style)
            ]]
            
            # 기둥과 보를 하나의 리스트로 취합
            sheet_members = []
            for col in sheet_cols:
                txt = col.get("text", "").upper().strip()
                spec = get_spec_detail(sheet_id, txt)
                h_mm = float(col.get("height", 0))
                uw = get_hbeam_unit_weight(spec, txt, is_column=True)
                weight = (h_mm / 1000.0) * uw / 1000.0  # ton 단위
                sheet_members.append({
                    "type": "기둥",
                    "text": txt,
                    "spec": spec,
                    "val": h_mm,
                    "uw": uw,
                    "weight": weight,
                    "x": float(col.get("cx", 0)),
                    "y": float(col.get("cy", 0))
                })
                
            for beam in sheet_beams:
                txt = beam.get("text", "").upper().strip()
                spec = get_spec_detail(sheet_id, txt)
                l_mm = float(beam.get("length", 0))
                uw = get_hbeam_unit_weight(spec, txt, is_column=False)
                weight = (l_mm / 1000.0) * uw / 1000.0  # ton 단위
                
                # 보의 시작점 좌표
                start_pt = beam.get("start", [0, 0])
                sheet_members.append({
                    "type": "보",
                    "text": txt,
                    "spec": spec,
                    "val": l_mm,
                    "uw": uw,
                    "weight": weight,
                    "x": float(start_pt[0]),
                    "y": float(start_pt[1])
                })
                
            # 정렬 기준: 부호 알파벳순 -> x좌표 오름차순 -> y좌표 내림차순 (큰 순서)
            sheet_members.sort(key=lambda m: (
                m["text"],
                m["x"],
                -m["y"]
            ))
            
            for m in sheet_members:
                details_table_data.append([
                    Paragraph(m["type"], normal_style),
                    Paragraph(m["text"], normal_style),
                    Paragraph(m["spec"], normal_style),
                    Paragraph(f"{m['val']:,.0f}", normal_style),
                    Paragraph(f"{m['uw'] / 1000.0:,.4f}", normal_style),
                    Paragraph(f"{m['weight']:,.3f}", normal_style),
                    Paragraph(f"{m['x']:,.1f}", normal_style),
                    Paragraph(f"{m['y']:,.1f}", normal_style)
                ])
                
            if len(sheet_members) > 0:
                t_details = Table(details_table_data, colWidths=[45, 50, 123, 65, 55, 65, 60, 60])
                t_details.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4A607A')),
                    ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                    ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
                    ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')]),
                ]))
                story.append(t_details)
                
            story.append(PageBreak())
        
        if story and isinstance(story[-1], PageBreak):
            story.pop()
            
        doc.build(story)
    except Exception as pdf_err:
        print(f"Failed to generate PDF: {pdf_err}")
        import traceback
        traceback.print_exc()
        pdf_filename = None

    return {
        "status": "success",
        "excel_file": saved_excel_filename,
        "pdf_file": pdf_filename,
        "marked_dxf_file": marked_dxf_filename
    }

# API 4: 파일 다운로드
@app.get("/api/download/{filename}")
async def download_file(filename: str):
    if "_takeoff" in filename or "_marked.dxf" in filename:
        file_path = os.path.join(OUTPUT_DIR, filename)
    else:
        file_path = os.path.join(DATA_DIR, filename)
        
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Requested file not found.")
    return FileResponse(file_path, filename=filename)



class AiHelperRequest(BaseModel):
    filename: str
    sheet_id: str
    bbox: List[float]
    existing_data: Optional[List[dict]] = None

@app.post("/api/ai_analyze_columns")
async def ai_analyze_columns(request: AiHelperRequest):
    if not OPENROUTER_API_KEY:
        raise HTTPException(status_code=500, detail="OpenRouter API key is not configured. Please set OPENROUTER_API_KEY in your env.")

    file_path = os.path.join(DATA_DIR, request.filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Target DXF file not found.")

    # 1. DXF 정보 추출
    try:
        dxf_data = extract_dxf_data_for_ai(file_path, request.bbox)
        if "error" in dxf_data:
            raise HTTPException(status_code=500, detail=dxf_data["error"])
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to extract DXF data: {str(e)}")

    # 2. rules_reference.md 로드
    rules_ref = ""
    rules_path = os.path.join(BASE_DIR, "rules_reference.md")
    if os.path.exists(rules_path):
        try:
            with open(rules_path, "r", encoding="utf-8") as f:
                rules_ref = f.read()
        except Exception as e:
            print(f"Failed to read rules reference: {e}")

    # 3. 프롬프트 구성
    sys_prompt = (
        "너는 건축 구조 도면의 DXF 텍스트와 형상 정보를 분석하여 H형강 기둥(Column)을 정확히 분류하고 수량과 중심 좌표를 산출하는 AI 엔지니어다.\n"
        "다음의 [규칙 참고서(Rules Reference)]를 철저히 준수하여 도면에 매핑되지 않았거나 잘못 매핑된 기둥을 찾고, 수정된 기둥 리스트를 제안해야 한다.\n"
        "특히 [규칙 참고서]의 '6. 도면 작성 편차에 대한 유연한 AI 분석 및 보완 규칙'에 의거하여, 도면 작성자마다 다르게 작성하는 규칙 불일치(텍스트 배치 편차, 단면 폴리선 불완전성 등)에 유연하게 대응해야 한다. 기존 규칙 기반 알고리즘의 분석 데이터(existing_data)는 오직 1차 참고용으로만 활용하며, 이를 맹신하지 말고 스스로 공간적 관계와 구조적 배치를 고려하여 오탐지는 제거하고 누락된 기둥은 발굴해내라.\n\n"
        f"[규칙 참고서]\n{rules_ref}\n"
    )

    user_content = (
        f"현재 분석 중인 시트 ID: {request.sheet_id}\n"
        f"현재 도면 영역 BBox: {request.bbox}\n"
        f"추출된 DXF 데이터 (텍스트 및 폴리선 후보): {json.dumps(dxf_data, ensure_ascii=False, indent=2)}\n"
        f"기존 규칙 기반 알고리즘이 분석한 1차 기둥 리스트: {json.dumps(request.existing_data, ensure_ascii=False, indent=2)}\n\n"
        "분석 미션:\n"
        "1. 기존 규칙 기반 1차 기둥 리스트(existing_data)를 맹신하지 말고, DXF 텍스트(texts)와 기둥 단면 후보(column_candidates)의 위치를 교차 비교하여 잘못 검출된 매칭 오류는 적극 제거하거나 보정하라.\n"
        "2. 도면 작성자가 기둥 텍스트를 기둥 단면 기하로부터 멀리(3,600mm 초과) 배치했을 가능성(Offset)을 염두에 두라. 그리드 축 정렬이나 대칭적인 기둥 간격 흐름을 파악하여, 거리가 기준을 다소 초과하더라도 구조적 맥락상 확실한 기둥이라면 올바르게 매칭을 승인하라.\n"
        "3. 기둥 단면이 닫힌 폴리선이 아니거나 분절된 다수의 라인으로 엉성하게 그려진 경우라도, H빔 단면 모양(가로세로 100~1500mm 스케일)을 이루고 있다면 적극적으로 기둥 중심 좌표를 산출하여 기둥으로 매칭하라.\n"
        "4. 규칙 기반 알고리즘이 누락한(놓친) 기둥 단면이 있다면, DXF 데이터 상에서 적절한 기둥 텍스트를 찾아 새로 기둥으로 추가하라.\n"
        "5. **중요**: 반드시 제공된 DXF 데이터 내 'schedule_table' 리스트에 존재하는 유효한 기둥 부호(symbol)만을 기둥 분석 및 매칭 결과에 포함해야 한다. 만약 어떤 기둥 부호(예: SC1)가 도면 텍스트에 나타나 있더라도, 'schedule_table' 리스트에 존재하지 않는다면(예: B로 시작하는 비H빔 규격 등을 가져 제외된 부호 등) 기둥 매칭 제안 목록(ai_columns)에서 절대적으로 배제(제외)해야 한다. 존재하지 않는 임의의 부호를 새로 만들어내지 마라.\n"
        "6. 분석 결과는 반드시 JSON 포맷으로만 응답하라. 마크다운 기호(```json 등)나 기타 부연설명은 절대 포함하지 말고 순수 JSON만 응답해야 한다.\n\n"
        "7. **중요**: 반환하는 cx, cy 좌표는 반드시 'column_candidates' 배열에 있는 기둥 단면(I자 형상 폴리선)의 중심 좌표여야 한다. 기둥 부호 텍스트(SC2, MC1 등)의 좌표를 반환하면 안 된다. 기둥 부호 텍스트는 기둥 단면과 떨어져 있을 수 있으므로, 텍스트 좌표가 아닌 실제 기둥 단면의 중심 좌표를 사용하라.\n\n"
        "반환 형식:\n"
        "{\n"
        '  "ai_columns": [\n'
        "    {\n"
        '      "symbol": "기둥부호 (예: MC1)",\n'
        '      "cx": 기둥중심 X좌표 (숫자),\n'
        '      "cy": 기둥중심 Y좌표 (숫자),\n'
        '      "confidence": 매칭 신뢰도 (0.0~1.0 사이 실수),\n'
        '      "reason": "해당 기둥을 판정한 근거 (예: 기존 룰베이스 누락 복원, 오프셋 텍스트 매칭 등)"\n'
        "    }\n"
        "  ],\n"
        '  "summary": "총 12개 기둥 발견",\n'
        f'  "model_used": "{OPENROUTER_MODEL}"\n'
        "}"
    )

    # 4. OpenRouter API 호출
    headers = {
        "Authorization": f"Bearer {OPENROUTER_API_KEY}",
        "Content-Type": "application/json"
    }
    payload = {
        "model": OPENROUTER_MODEL,
        "messages": [
            {"role": "system", "content": sys_prompt},
            {"role": "user", "content": user_content}
        ],
        "response_format": {"type": "json_object"},
        "max_tokens": 16384
    }
    payload = sanitize_surrogates(payload)

    try:
        async with httpx.AsyncClient(timeout=60.0) as client:
            response = await client.post("https://openrouter.ai/api/v1/chat/completions", headers=headers, json=payload)
            if response.status_code != 200:
                raise HTTPException(status_code=response.status_code, detail=f"OpenRouter API error: {response.text}")
            
            res_json = response.json()
            ai_message = res_json['choices'][0]['message']['content'].strip()
            
            try:
                ai_data = json.loads(ai_message)
                return ai_data
            except Exception as e:
                cleaned_message = ai_message.replace("```json", "").replace("```", "").strip()
                return json.loads(cleaned_message)

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"AI column analysis failed: {str(e)}")

@app.post("/api/ai_analyze_beams")
async def ai_analyze_beams(request: AiHelperRequest):
    if not OPENROUTER_API_KEY:
        raise HTTPException(status_code=500, detail="OpenRouter API key is not configured. Please set OPENROUTER_API_KEY in your env.")

    file_path = os.path.join(DATA_DIR, request.filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Target DXF file not found.")

    # 1. DXF 정보 추출
    try:
        dxf_data = extract_dxf_data_for_ai(file_path, request.bbox)
        if "error" in dxf_data:
            raise HTTPException(status_code=500, detail=dxf_data["error"])
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to extract DXF data: {str(e)}")

    # 2. rules_reference.md 로드
    rules_ref = ""
    rules_path = os.path.join(BASE_DIR, "rules_reference.md")
    if os.path.exists(rules_path):
        try:
            with open(rules_path, "r", encoding="utf-8") as f:
                rules_ref = f.read()
        except Exception as e:
            print(f"Failed to read rules reference: {e}")

    # 3. 프롬프트 구성
    sys_prompt = (
        "너는 건축 구조 도면의 DXF 텍스트와 형상 정보를 분석하여 H형강 보(Beam/Girder)를 정확히 분류하고 시작/끝 좌표를 산출하는 AI 엔지니어다.\n"
        "다음의 [규칙 참고서(Rules Reference)]를 철저히 준수하여 도면에 매핑되지 않았거나 잘못 매핑된 보를 찾고, 수정된 보 리스트를 제안해야 한다.\n"
        "특히 [규칙 참고서]의 '6. 도면 작성 편차에 대한 유연한 AI 분석 및 보완 규칙'에 의거하여, 도면 작성자마다 다르게 작성하는 규칙 불일치(텍스트 배치 편차, 선분 분절 등)에 유연하게 대응해야 한다. 기존 규칙 기반 알고리즘의 분석 데이터(existing_data)는 오직 1차 참고용으로만 활용하며, 이를 맹신하지 말고 스스로 공간적 관계와 구조적 배치를 고려하여 오탐지는 제거하고 누락된 보선은 발굴해내라.\n\n"
        f"[규칙 참고서]\n{rules_ref}\n"
    )

    user_content = (
        f"현재 분석 중인 시트 ID: {request.sheet_id}\n"
        f"현재 도면 영역 BBox: {request.bbox}\n"
        f"추출된 DXF 데이터 (텍스트 및 선분 후보): {json.dumps(dxf_data, ensure_ascii=False, indent=2)}\n"
        f"기존 규칙 기반 알고리즘이 분석한 1차 보 리스트: {json.dumps(request.existing_data, ensure_ascii=False, indent=2)}\n\n"
        "분석 미션:\n"
        "1. 기존 규칙 기반 1차 보 리스트(existing_data)를 맹신하지 말고, DXF 텍스트(texts)와 보 기하선 후보(beam_candidates)의 위치 및 각도를 교차 비교하여 잘못 검출된 매칭 오류는 적극 제거하거나 보정하라.\n"
        "2. 도면 작성자가 보 부호 텍스트를 보 기하선으로부터 멀리(4,000mm 초과) 배치하거나 텍스트 정렬 각도가 약간 틀어졌을 가능성을 염두에 두라. 구조적 연결성 및 그리드 맥락상 확실하게 일치하는 쌍이라면 거리가 기준을 다소 초과하더라도 유연하게 매칭을 승인하라.\n"
        "3. 보 기하선이 하나의 단일 선분으로 정의되지 않고 여러 조각(LINE)으로 분절되어 있어도, 연장선 상에서 하나의 보 흐름을 이루고 있다면 이를 결합하고 최종 시작점과 끝점을 추론하여 매칭하라.\n"
        "4. 규칙 기반 알고리즘이 누락한(놓친) 보가 존재한다면, DXF 데이터 상에서 적절한 보 텍스트와 기하선을 찾아 새로 보로 추가하라.\n"
        "5. **중요**: 반드시 제공된 DXF 데이터 내 'schedule_table' 리스트에 존재하는 유효한 보 부호(symbol)만을 보 분석 및 매칭 결과에 포함해야 한다. 만약 어떤 보 부호가 도면 텍스트에 나타나 있더라도, 'schedule_table' 리스트에 존재하지 않는다면 보 매칭 제안 목록(ai_beams)에서 절대적으로 배제(제외)해야 한다. 존재하지 않는 임의의 부호를 새로 만들어내지 마라.\n"
        "6. 분석 결과는 반드시 JSON 포맷으로만 응답하라. 마크다운 기호(```json 등)나 기타 부연설명은 절대 포함하지 말고 순수 JSON만 응답해야 한다.\n\n"
        "반환 형식:\n"
        "{\n"
        '  "ai_beams": [\n'
        "    {\n"
        '      "symbol": "보부호 (예: G1)",\n'
        '      "start": [시작점X, 시작점Y],\n'
        '      "end": [끝점X, 끝점Y],\n'
        '      "length": 보의길이(mm, 숫자),\n'
        '      "confidence": 매칭 신뢰도 (0.0~1.0 사이 실수),\n'
        '      "reason": "해당 보를 판정한 근거 (예: 분절된 보선 병합, 오프셋 텍스트 매칭 등)"\n'
        "    }\n"
        "  ],\n"
        '  "summary": "총 24개 보 발견",\n'
        f'  "model_used": "{OPENROUTER_MODEL}"\n'
        "}"
    )

    # 4. OpenRouter API 호출
    headers = {
        "Authorization": f"Bearer {OPENROUTER_API_KEY}",
        "Content-Type": "application/json"
    }
    payload = {
        "model": OPENROUTER_MODEL,
        "messages": [
            {"role": "system", "content": sys_prompt},
            {"role": "user", "content": user_content}
        ],
        "response_format": {"type": "json_object"},
        "max_tokens": 16384
    }
    payload = sanitize_surrogates(payload)

    try:
        async with httpx.AsyncClient(timeout=120.0) as client:  # 60 -> 120
            response = await client.post("https://openrouter.ai/api/v1/chat/completions", headers=headers, json=payload)
            if response.status_code != 200:
                raise HTTPException(status_code=response.status_code, detail=f"OpenRouter API error: {response.text}")
            
            res_json = response.json()
            ai_message = res_json['choices'][0]['message']['content'].strip()
            
            # 디버깅: AI 원본 응답 로깅
            print(f"[DEBUG AI BEAM] 원본 응답 (처음 500자): {ai_message[:500]}")
            
            try:
                ai_data = json.loads(ai_message)
                return ai_data
            except Exception as e:
                print(f"[DEBUG AI BEAM] JSON 파싱 실패: {str(e)}")
                cleaned_message = ai_message.replace("```json", "").replace("```", "").strip()
                print(f"[DEBUG AI BEAM] 정리 후 응답 (처음 500자): {cleaned_message[:500]}")
                return json.loads(cleaned_message)

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"AI beam analysis failed: {str(e)}")

class AiHeightRequest(BaseModel):
    filename: str
    sheet_id: str
    ref_sheet_id: Optional[str] = None
    bbox: List[float]
    texts: Optional[List[Dict[str, Any]]] = None

@app.post("/api/ai_recommend_height")
async def ai_recommend_height(request: AiHeightRequest):
    if not OPENROUTER_API_KEY:
        raise HTTPException(status_code=500, detail="OpenRouter API key is not configured. Please set OPENROUTER_API_KEY in your env.")

    file_path = os.path.join(DATA_DIR, request.filename)
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Target DXF file not found.")

    target_bbox = request.bbox
    sheet_name_to_log = request.sheet_id

    if request.ref_sheet_id:
        meta_path = file_path + ".meta.json"
        if os.path.exists(meta_path):
            try:
                with open(meta_path, "r", encoding="utf-8") as f:
                    meta_data = json.load(f)
                    for s in meta_data.get("sheets", []):
                        if s["id"] == request.ref_sheet_id:
                            target_bbox = s["bbox"]
                            sheet_name_to_log = s["name"]
                            break
            except Exception as e:
                print(f"Failed to load meta file: {e}")

    try:
        if request.texts is not None:
            # 백엔드 자체 수직선 추출기와 크로스 체크하여 치수선 타입(dimension)의 수직선 여부를 실시간 검증
            real_verticals = extract_height_texts_for_ai(file_path, target_bbox)
            if isinstance(real_verticals, dict) and "error" in real_verticals:
                real_verticals = []
            
            vertical_coords = set()
            for rv in real_verticals:
                vertical_coords.add((round(rv['x']/10)*10, round(rv['y']/10)*10))
            
            height_texts = []
            for t in request.texts:
                t_direction = t.get('direction', 'unknown')
                t_type = t.get('type', '')
                
                if t_direction == 'horizontal':
                    continue
                
                if t_type == 'dimension':
                    # 좌표 기반 검증
                    tx = float(t.get('x', 0))
                    ty = float(t.get('y', 0))
                    key = (round(tx/10)*10, round(ty/10)*10)
                    if key in vertical_coords:
                        t['direction'] = 'vertical'
                        height_texts.append(t)
                else:
                    # 일반 텍스트는 그대로 전달
                    height_texts.append(t)
        else:
            height_texts = extract_height_texts_for_ai(file_path, target_bbox)
            if isinstance(height_texts, dict) and "error" in height_texts:
                raise HTTPException(status_code=500, detail=height_texts["error"])
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to extract height texts: {str(e)}")

    if not height_texts:
        return {"recommended_height": 0, "reason": "도면 영역 내에서 기둥 높이 후보 수치 텍스트를 감지할 수 없습니다."}

    rules_ref = ""
    rules_path = os.path.join(BASE_DIR, "rules_reference.md")
    if os.path.exists(rules_path):
        try:
            with open(rules_path, "r", encoding="utf-8") as f:
                rules_ref = f.read()
        except Exception as e:
            print(f"Failed to read rules reference: {e}")

    sys_prompt = (
        "너는 건축 구조 도면 및 단면도/입면도를 정밀하게 분석하여 기둥 층고(높이)를 추천하는 AI 엔지니어다.\n"
        "다음의 [규칙 참고서]를 준수하고, 도면에 있는 수치 텍스트 데이터에서 최적의 기둥 높이를 판단해야 한다.\n\n"
        f"[규칙 참고서]\n{rules_ref}\n\n"
        "[중요] 모든 응답은 반드시 한국어(한글)로만 작성해야 한다. 영어 사용 금지.\n"
    )

    user_content = (
        f"현재 분석 중인 도면 영역 ({sheet_name_to_log})에서 추출된 높이/수치 후보 텍스트 데이터:\n"
        f"{json.dumps(height_texts, ensure_ascii=False, indent=2)}\n\n"
        "분석 미션:\n"
        "1. 도면 데이터에 나타난 치수선 값(type=dimension, direction=vertical)이나 층고 텍스트(예: 3000, 3600, 3800, 4200, 4500, 5000 등) 중 **가장 크고 유효한(제일 긴) 기둥 높이값**을 찾으시오.\n"
        "2. 일반적으로 층고 수치는 밀리미터(mm) 단위로 표현되며, 대개 2000mm ~ 15000mm 사이의 값을 가집니다.\n"
        "3. 만약 텍스트 중에 '1FL', '2FL', '3FL' 등의 레벨 차이 정보가 있고 이를 통해 산출 가능한 가장 큰 높이가 있다면 그것을 선택할 수도 있습니다.\n"
        "4. 분석 결과는 반드시 JSON 포맷으로만 응답하며, 어떠한 마크다운 기호나 추가 설명도 배제하시오.\n"
        "5. reason 필드는 반드시 한국어(한글)로만 작성하시오. 영어 사용 절대 금지.\n\n"
        "**중요 규칙**: 기둥 높이는 반드시 수직선(Vertical Line)의 치수만 사용해야 합니다. 수평선(Horizontal Line)의 길이는 기둥 높이로 인정하지 않습니다. direction='vertical'로 표시된 치수선만 기둥 높이 후보로 고려하십시오.\n\n"
        "반환 형식:\n"
        "{\n"
        '  "recommended_height": 제일 긴 기둥 높이값 (정수형 mm 단위, 예: 4200),\n'
        '  "reason": "해당 높이를 도면에서 어떻게 도출했는지에 대한 자세한 한글 근거 설명 (반드시 한국어로)"\n'
        "}"
    )

    headers = {
        "Authorization": f"Bearer {OPENROUTER_API_KEY}",
        "Content-Type": "application/json"
    }
    payload = {
        "model": OPENROUTER_MODEL,
        "messages": [
            {"role": "system", "content": sys_prompt},
            {"role": "user", "content": user_content}
        ],
        "response_format": {"type": "json_object"},
        "max_tokens": 16384
    }
    payload = sanitize_surrogates(payload)

    try:
        async with httpx.AsyncClient(timeout=60.0) as client:
            response = await client.post("https://openrouter.ai/api/v1/chat/completions", headers=headers, json=payload)
            if response.status_code != 200:
                raise HTTPException(status_code=response.status_code, detail=f"OpenRouter API error: {response.text}")
            
            res_json = response.json()
            ai_message = res_json['choices'][0]['message']['content'].strip()
            
            try:
                ai_data = json.loads(ai_message)
                return ai_data
            except Exception:
                cleaned_message = ai_message.replace("```json", "").replace("```", "").strip()
                return json.loads(cleaned_message)

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"AI height recommendation failed: {str(e)}")



@app.get("/", response_class=HTMLResponse)
async def read_index():
    index_path = os.path.join(BASE_DIR, "index.html")
    if not os.path.exists(index_path):
        return HTMLResponse(content="<h3>index.html not found. Please wait until front-end is generated.</h3>", status_code=404)
    with open(index_path, "r", encoding="utf-8") as f:
        html_content = f.read()
    return HTMLResponse(content=html_content)


if __name__ == "__main__":
#    uvicorn.run("server:app", host="127.0.0.1", port=8000, reload=True)
    uvicorn.run("server:app", host="127.0.0.1", port=8000)
